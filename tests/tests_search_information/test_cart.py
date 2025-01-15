import os
import allure
from model.pages.search_information_page import SearchInformationPage
from selenium import webdriver
from openpyxl import load_workbook
from resources.script_os import TMP_DIR


@allure.tag('web')
@allure.feature("Загрузка файла из корзины")
@allure.label("owner", "Ярослав Гусев")
@allure.description("Тест добавления файлов в корзину и проверки заказа")
@allure.link("https://www.cft.ru/", name="Testing")
def test_cart():
    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": TMP_DIR,
        "download.prompt_for_download": False,
    }
    options.add_experimental_option("prefs", prefs)

    search = SearchInformationPage()

    with allure.step('Открывает сайт "https://www.cft.ru/"'):
        search.open()

    with allure.step('Открываем "Каталоги решений и продуктов"'):
        search.switch_page('Каталоги решений')

    with allure.step('Открываем боковую панель"'):
        search.open_sidebar()

    with allure.step('Открываем вкладку "Готовые Решения"'):
        search.module_choice('Готовые Решения')

    with allure.step('Переходим на страницу "Универсальный банк"'):
        search.module_choice('Универсальный банк')

    with allure.step('Добавляем все в корзину"'):
        search.add_all_in_basket()

    with allure.step('Переходим в корзину'):
        search.user_dashboard_actions('Корзина')

    with allure.step('Добавляем зависимости, выгружаем Exel файл и очищаем корзину'):
        search.process('Добавить зависимости')
        search.process('Выгрузить в Excel')
        search.process('Очистить корзину')


@allure.tag('web')
@allure.feature("Проверка загруженного файла")
@allure.label("owner", "Ярослав Гусев")
@allure.description("Тест проверки скачанного файла из корзины")
@allure.link("https://www.cft.ru/", name="Testing")
def test_downloaded_file():
    search = SearchInformationPage()

    with allure.step('Проверяем наличие скачанных файлов'):
        files = os.listdir(TMP_DIR)
        if len(files) > 0:
            for file in files:
                file_xlsx = os.path.join(TMP_DIR, file)
                try:
                    workbook = load_workbook(file_xlsx)
                    sheet = workbook.active
                    text_xlsx = ''
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value is None:
                                continue
                            text_xlsx += f'{cell.value} \n'
                    print('Файл существует и содержит информацию.')
                    allure.attach(text_xlsx, name=f'Content of {file}', attachment_type=allure.attachment_type.TEXT)
                except Exception as e:
                    print(f"Ошибка при чтении файла {file_xlsx}: {e}")
                    allure.attach(str(e), name=f'Error reading {file}', attachment_type=allure.attachment_type.TEXT)
        else:
            print('Нет папки, либо папка пуста!')
            allure.attach('No files found or directory is empty', name='Directory Check',
                          attachment_type=allure.attachment_type.TEXT)

    with allure.step('Удаляем временные файлы'):
        search.del_temporary_folder()
